import csv
import json
from pathlib import Path

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from models.decisions import Decision
from models.evidence import Evidence
from models.migration import PipelineResult, RiskFinding


class ReportBundleWriter:
    def write(self, result: PipelineResult, directory: Path) -> tuple[Path, ...]:
        directory.mkdir(parents=True, exist_ok=True)
        workbook = directory / f"migration-report-{result.run_id}.xlsx"
        contacts = directory / f"contact-import-candidates-{result.run_id}.csv"
        entitlements = directory / f"entitlement-import-candidates-{result.run_id}.csv"
        manual_review = directory / f"manual-review-{result.run_id}.csv"
        manifest = directory / f"run-manifest-{result.run_id}.json"
        self._workbook(result, workbook)
        self._csv(contacts, result.prepared_dataset.contact_rows, [
            "Email", "Name", "Phone", "Source File", "Source Row", "Import Eligible",
            "Activation Authorized",
        ])
        self._csv(entitlements, result.prepared_dataset.rows, [
            "Email", "Entitlement Key", "Decision Status", "Evidence Count",
            "Activation Authorized",
        ])
        self._csv(manual_review, result.prepared_dataset.manual_review_rows, [
            "Subject", "Decision Type", "Status", "Rationale", "Conflicts",
            "Evidence Count", "Approved", "Reviewer Notes",
        ])
        manifest.write_text(json.dumps({
            "run_id": result.run_id, "release_status": result.release_status,
            "sources": [item.model_dump(mode="json") for item in result.source_manifests],
            "counts": {
                "contact_candidates": len(result.prepared_dataset.contact_rows),
                "entitlement_candidates": len(result.prepared_dataset.rows),
                "manual_review": len(result.prepared_dataset.manual_review_rows),
            },
            "artifacts": [workbook.name, contacts.name, entitlements.name, manual_review.name],
        }, indent=2) + "\n", encoding="utf-8")
        return tuple(path.resolve() for path in (
            workbook, contacts, entitlements, manual_review, manifest
        ))

    @staticmethod
    def _csv(path: Path, rows: tuple[dict[str, object], ...], fields: list[str]) -> None:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)

    def _workbook(self, result: PipelineResult, path: Path) -> None:
        book = Workbook()
        summary = book.active
        summary.title = "Executive Summary"
        summary.append(["Metric", "Value"])
        for row in (
            ("Run ID", result.run_id), ("Release Status", result.release_status),
            ("Source Files", len(result.source_manifests)),
            ("Normalized Records", len(result.normalized_records)),
            ("Duplicate Candidates", len(result.duplicate_decisions)),
            ("Entitlement Decisions", len(result.entitlement_decisions)),
            ("Risk Findings", len(result.risks)),
            ("Candidate Import Rows", len(result.prepared_dataset.rows)),
            ("Contact Import Candidates", len(result.prepared_dataset.contact_rows)),
            ("Manual Review Rows", len(result.prepared_dataset.manual_review_rows)),
            ("Activation Authorized", "NO"),
        ):
            summary.append(row)
        self._style(summary)
        self._decisions(book.create_sheet("Duplicate Review"), result.duplicate_decisions)
        self._decisions(book.create_sheet("Entitlement Ledger"), result.entitlement_decisions)
        self._risks(book.create_sheet("QA & Exceptions"), result.risks)
        self._records(book.create_sheet("Normalized Records"), result)
        self._sources(book.create_sheet("Source Register"), result)
        book.save(path)

    def _records(self, sheet: Worksheet, result: PipelineResult) -> None:
        sheet.append([
            "Record Type", "Normalized Email", "Normalized Name", "Normalized Phone",
            "Source File", "Source Row", "Evidence",
        ])
        for item in result.normalized_records:
            sheet.append([
                item.record_type, item.normalized_email, item.normalized_name,
                item.normalized_phone, item.source_file, item.source_row,
                self._evidence(item.evidence),
            ])
        self._style(sheet)

    @staticmethod
    def _sources(sheet: Worksheet, result: PipelineResult) -> None:
        sheet.append([
            "Source File", "Export Type", "SHA256", "Rows Seen", "Rows Accepted",
            "Rows Quarantined", "Encoding", "Delimiter", "Columns",
        ])
        for item in result.source_manifests:
            sheet.append([
                item.source_file, item.export_type.value, item.sha256, item.rows_seen,
                item.rows_accepted, item.rows_quarantined, item.encoding, item.delimiter,
                "; ".join(item.columns),
            ])
        ReportBundleWriter._style(sheet)

    def _decisions(self, sheet: Worksheet, decisions: tuple[Decision, ...]) -> None:
        sheet.append(["Subject", "Decision Type", "Status", "Rationale", "Evidence", "Conflicts"])
        for item in decisions:
            sheet.append([item.subject_id, item.decision_type, item.status.value, item.rationale,
                          self._evidence(item.evidence), "; ".join(item.conflicts)])
        self._style(sheet)

    def _risks(self, sheet: Worksheet, risks: tuple[RiskFinding, ...]) -> None:
        sheet.append(["Severity", "Code", "Subject", "Summary", "Recommended Action", "Evidence"])
        for item in risks:
            sheet.append([item.severity.value, item.code, item.subject_id, item.summary,
                          item.recommended_action, self._evidence(item.evidence)])
        self._style(sheet)

    @staticmethod
    def _evidence(items: tuple[Evidence, ...]) -> str:
        return "; ".join(f"{x.source_file}:{x.source_row}:{x.source_field}" for x in items)

    @staticmethod
    def _style(sheet: Worksheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(
                max(len(str(cell.value or "")) for cell in column) + 2, 60
            )
